#!/usr/bin/env python3
"""Generate MachineFit catalog upload Excel template."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path("/workspace")
OUT = ROOT / "docs" / "templates" / "MachineFit_카탈로그_업로드_양식.xlsx"
OUT_COPY = ROOT / "database" / "catalog" / "templates" / "MachineFit_카탈로그_업로드_양식.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="111827")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
REQ_FILL = PatternFill("solid", fgColor="FEF3C7")
OPT_FILL = PatternFill("solid", fgColor="ECFDF5")
SAMPLE_FILL = PatternFill("solid", fgColor="EFF6FF")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def style_header(ws, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(cols)}1"


def autosize(ws, widths: dict[int, int]):
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_rows(ws, rows: list[list], start_row=2, fill=None):
    for r_i, row in enumerate(rows):
        for c_i, value in enumerate(row, start=1):
            cell = ws.cell(start_row + r_i, c_i, value)
            cell.alignment = WRAP
            cell.border = THIN
            if fill:
                cell.fill = fill


def sheet_guide(wb: Workbook):
    ws = wb.active
    ws.title = "00_작성안내"
    lines = [
        ["MachineFit 카탈로그 업로드 엑셀 양식"],
        [""],
        ["목적"],
        ["브랜드 / 브랜드별 기구 / 사진 / 주의사항 / 운동팁 등 서비스에 표시(또는 표시 예정)되는 항목을 엑셀로 작성하기 위한 샘플 양식입니다."],
        [""],
        ["시트 구성"],
        ["01_필드사전 — 서비스에 쓰이는 모든 항목 목록 (표시중 / 데이터만 / 관리자)"],
        ["02_브랜드 — 브랜드 목록·상세·관리자에 표시되는 항목"],
        ["03_머신 — 머신 검색/상세/추천에 쓰이는 기본 정보 + 가이드 문장"],
        ["04_머신이미지 — 대표 이미지·썸네일·갤러리(복수 가능)"],
        ["05_추천설정_팁주의 — 추천 결과 화면의 시트/패드/중량/횟수 + 운동팁·주의사항(음성 안내 포함)"],
        ["06_헬스장보유머신 — 헬스장 상세에 표시되는 보유 머신 정보(참고)"],
        ["07_작성예시 — 한 줄 샘플(해머 스트렝스 / 레그 컬)"],
        [""],
        ["작성 규칙"],
        ["1) 노란 헤더/열 = 필수, 초록 = 선택"],
        ["2) 다국어는 ko / en 을 우선 작성 (ja/zh는 선택)"],
        ["3) 팁·주의·사용법 등은 한 셀에 여러 줄(Alt+Enter)로 작성. 줄마다 한 문장. TTS로 읽히므로 자연스러운 문장/쉼표 사용"],
        ["4) 이미지 파일명은 assets 폴더에 넣을 파일명과 동일하게 (예: cy_leg_curl.jpg). URL이 있으면 이미지URL 열에 작성"],
        ["5) machineCode / brandCode 는 영문 대문자+언더스코어 (예: CY_LEG_CURL, CYBEX)"],
        ["6) 운동부위(muscleGroup): chest, back, legs, shoulders, biceps, triceps, core, full_body"],
        ["7) 머신종류(machineType): plate_loaded, selectorized, cable, smith, free_weight, bodyweight"],
        ["8) 추천경험레벨: beginner, intermediate, advanced"],
        ["9) 현재 UI에 바로 안 보이는 항목도 포함했습니다(가이드 문장, 유튜브 등). 나중에 화면 연결용으로 미리 채워 주세요."],
        [""],
        ["현재 화면 표시 요약"],
        ["표시중 — 브랜드명/소개/로고/사이트, 머신명/브랜드명/부위/종류/대표이미지, 추천 시트·백패드·발·핸들·가동범위·중량·횟수, 운동팁·주의, 헬스장 보유머신"],
        ["데이터만(미표시) — 머신 description, howTo, beginnerTips, recommendedExperience, 이미지 갤러리/alt, YouTube"],
        [""],
        ["문의/확장"],
        ["브랜드 추가 시 02_브랜드 + 03_머신 + 04_머신이미지 + 05_추천설정_팁주의 를 함께 채우면 됩니다."],
    ]
    for i, row in enumerate(lines, start=1):
        ws.cell(i, 1, row[0] if row else "")
        ws.cell(i, 1).alignment = WRAP
        if i == 1:
            ws.cell(i, 1).font = Font(bold=True, size=16, color="111827")
        elif row and row[0] in {"목적", "시트 구성", "작성 규칙", "현재 화면 표시 요약", "문의/확장"}:
            ws.cell(i, 1).font = Font(bold=True, size=12, color="065F46")
    ws.column_dimensions["A"].width = 120


def sheet_dictionary(wb: Workbook):
    ws = wb.create_sheet("01_필드사전")
    headers = [
        "구분",
        "필드키(API/DB)",
        "한글표시명",
        "영문표시명",
        "표시화면",
        "표시상태",
        "필수여부",
        "데이터형식",
        "작성시트",
        "비고",
    ]
    ws.append(headers)
    style_header(ws, len(headers))
    rows = [
        # Brand
        ["브랜드", "brandCode", "브랜드코드", "Brand Code", "브랜드목록/상세/관리자", "표시중", "필수", "문자열 UPPER_SNAKE", "02_브랜드", "예: CYBEX"],
        ["브랜드", "name.ko", "브랜드명(한글)", "Brand Name KO", "브랜드목록/상세/필터칩", "표시중", "필수", "문자열", "02_브랜드", ""],
        ["브랜드", "name.en", "브랜드명(영문)", "Brand Name EN", "브랜드목록/상세/관리자", "표시중", "필수", "문자열", "02_브랜드", ""],
        ["브랜드", "name.ja", "브랜드명(일본어)", "Brand Name JA", "다국어", "선택표시", "선택", "문자열", "02_브랜드", ""],
        ["브랜드", "name.zh", "브랜드명(중국어)", "Brand Name ZH", "다국어", "선택표시", "선택", "문자열", "02_브랜드", ""],
        ["브랜드", "description.ko", "브랜드소개(한글)", "Brand Intro KO", "브랜드카드/상세", "표시중", "권장", "문장", "02_브랜드", "2~3문장"],
        ["브랜드", "description.en", "브랜드소개(영문)", "Brand Intro EN", "브랜드카드/상세", "표시중", "권장", "문장", "02_브랜드", ""],
        ["브랜드", "logoFile / logoUrl", "브랜드대표이미지", "Brand Logo", "브랜드카드/상세/관리자", "표시중", "권장", "파일명 또는 URL", "02_브랜드", "없으면 이니셜"],
        ["브랜드", "websiteUrl", "공식홈페이지", "Website", "브랜드상세", "표시중", "선택", "URL", "02_브랜드", ""],
        ["브랜드", "slug", "브랜드슬러그", "Brand Slug", "이미지경로", "내부용", "필수", "소문자_스네이크", "02_브랜드", "예: cybex"],
        ["브랜드", "isActive", "활성여부", "Active", "관리자", "표시중", "필수", "TRUE/FALSE", "02_브랜드", "기본 TRUE"],
        ["브랜드", "countryId", "국가ID", "Country ID", "-", "미표시", "선택", "UUID", "-", "현재 UI 없음"],
        # Machine core
        ["머신", "machineCode", "머신코드", "Machine Code", "관리자/라우팅/QR", "표시중", "필수", "문자열 UPPER_SNAKE", "03_머신", "예: CY_LEG_CURL"],
        ["머신", "brandCode", "브랜드코드", "Brand Code", "조인", "표시중", "필수", "브랜드코드와 동일", "03_머신", ""],
        ["머신", "name.ko", "머신한글명", "Machine Name KO", "검색목록/상세/즐겨찾기/최근/추천결과", "표시중", "필수", "문자열", "03_머신", ""],
        ["머신", "name.en", "머신영문명", "Machine Name EN", "검색목록/상세/관리자", "표시중", "필수", "문자열", "03_머신", ""],
        ["머신", "name.ja", "머신일본어명", "Machine Name JA", "다국어", "선택표시", "선택", "문자열", "03_머신", ""],
        ["머신", "name.zh", "머신중국어명", "Machine Name ZH", "다국어", "선택표시", "선택", "문자열", "03_머신", ""],
        ["머신", "brandName", "브랜드표시명", "Brand Name", "머신목록 부제", "표시중", "자동", "브랜드명 조인", "-", "작성불필요"],
        ["머신", "muscleGroup", "운동부위", "Muscle Group", "상세배지/필터/아이콘", "표시중", "필수", "enum", "03_머신", "chest/back/legs/shoulders/biceps/triceps/core/full_body"],
        ["머신", "machineType", "머신종류", "Machine Type", "상세배지", "표시중", "필수", "enum", "03_머신", "plate_loaded/selectorized/cable/smith/free_weight/bodyweight"],
        ["머신", "description.ko", "머신소개(한글)", "Machine Intro KO", "상세(예정)", "데이터만", "권장", "문장", "03_머신", "현재 API만"],
        ["머신", "description.en", "머신소개(영문)", "Machine Intro EN", "상세(예정)", "데이터만", "권장", "문장", "03_머신", ""],
        ["머신", "howTo.ko", "사용방법(한글)", "How To KO", "상세(예정)/TTS", "데이터만", "권장", "줄바꿈 문장목록", "03_머신", "10개 이상 권장"],
        ["머신", "howTo.en", "사용방법(영문)", "How To EN", "상세(예정)/TTS", "데이터만", "권장", "줄바꿈 문장목록", "03_머신", ""],
        ["머신", "warnings.ko", "주의사항(머신가이드)", "Warnings KO", "상세(예정)", "데이터만", "권장", "줄바꿈 문장목록", "03_머신", "추천화면 주의와 별도 컬럼"],
        ["머신", "warnings.en", "주의사항(영문)", "Warnings EN", "상세(예정)", "데이터만", "권장", "줄바꿈 문장목록", "03_머신", ""],
        ["머신", "tips.ko", "운동꿀팁(머신가이드)", "Tips KO", "상세(예정)", "데이터만", "권장", "줄바꿈 문장목록", "03_머신", ""],
        ["머신", "tips.en", "운동꿀팁(영문)", "Tips EN", "상세(예정)", "데이터만", "권장", "줄바꿈 문장목록", "03_머신", ""],
        ["머신", "beginnerTips.ko", "초보자팁(한글)", "Beginner Tips KO", "상세(예정)", "데이터만", "권장", "줄바꿈 문장목록", "03_머신", "10개 이상 권장"],
        ["머신", "beginnerTips.en", "초보자팁(영문)", "Beginner Tips EN", "상세(예정)", "데이터만", "권장", "줄바꿈 문장목록", "03_머신", ""],
        ["머신", "recommendedExperience", "추천경험레벨", "Recommended Experience", "상세(예정)", "데이터만", "권장", "beginner/intermediate/advanced", "03_머신", ""],
        ["머신", "hasSeat", "시트유무", "Has Seat", "추천매칭용", "내부용", "필수", "TRUE/FALSE", "03_머신", "화면 직접표시 없음"],
        ["머신", "hasBackPad", "백패드유무", "Has Back Pad", "추천매칭용", "내부용", "필수", "TRUE/FALSE", "03_머신", ""],
        ["머신", "hasFootPlate", "발판유무", "Has Foot Plate", "추천매칭용", "내부용", "필수", "TRUE/FALSE", "03_머신", ""],
        ["머신", "hasHandle", "핸들유무", "Has Handle", "추천매칭용", "내부용", "필수", "TRUE/FALSE", "03_머신", ""],
        ["머신", "romType", "가동범위타입", "ROM Type", "추천기본값", "내부용", "선택", "fixed/variable/최대", "03_머신", ""],
        ["머신", "isActive", "활성여부", "Active", "관리자", "표시중", "필수", "TRUE/FALSE", "03_머신", ""],
        # Images
        ["이미지", "primaryImageFile / primaryImageUrl", "대표이미지", "Primary Image", "검색목록/상세히어로/관리자", "표시중", "권장", "파일명 또는 URL", "04_머신이미지", "isPrimary=TRUE"],
        ["이미지", "thumbnailFile", "썸네일이미지", "Thumbnail", "목록(예정)", "데이터만", "선택", "파일명", "04_머신이미지", "없으면 대표이미지 사용"],
        ["이미지", "galleryImageFile", "추가이미지", "Gallery Image", "갤러리(예정)", "데이터만", "선택", "파일명/URL", "04_머신이미지", "여러 행 가능"],
        ["이미지", "altText.ko/en", "이미지대체텍스트", "Alt Text", "접근성", "데이터만", "선택", "문자열", "04_머신이미지", ""],
        ["이미지", "sortOrder", "정렬순서", "Sort Order", "갤러리", "데이터만", "선택", "정수", "04_머신이미지", "0부터"],
        ["이미지", "isPrimary", "대표여부", "Is Primary", "목록/상세", "표시중", "필수", "TRUE/FALSE", "04_머신이미지", "머신당 TRUE 1개"],
        # Recommendation settings display
        ["추천설정", "seatPosition", "시트 위치", "Seat Position", "추천결과/기록/비교", "표시중", "조건부", "정수", "05_추천설정_팁주의", "hasSeat=TRUE일 때"],
        ["추천설정", "backPadPosition", "백 패드", "Back Pad", "추천결과/기록/비교", "표시중", "조건부", "정수", "05_추천설정_팁주의", ""],
        ["추천설정", "footPosition", "발 위치", "Foot Position", "추천결과/기록", "표시중", "조건부", "정수", "05_추천설정_팁주의", ""],
        ["추천설정", "handlePosition", "핸들 위치", "Handle Position", "추천결과/기록/비교", "표시중", "조건부", "정수", "05_추천설정_팁주의", ""],
        ["추천설정", "romSetting", "가동 범위", "Range of Motion", "추천결과/기록/비교", "표시중", "조건부", "문자열", "05_추천설정_팁주의", "최대/부분/반/짧게/가변"],
        ["추천설정", "recommendedWeightKg", "추천 중량", "Recommended Weight", "추천결과/기록/비교", "표시중", "권장", "숫자(kg)", "05_추천설정_팁주의", ""],
        ["추천설정", "recommendedRepsMin", "추천횟수최소", "Reps Min", "추천결과", "표시중", "선택", "정수", "05_추천설정_팁주의", "없으면 목표기반 계산"],
        ["추천설정", "recommendedRepsMax", "추천횟수최대", "Reps Max", "추천결과", "표시중", "선택", "정수", "05_추천설정_팁주의", ""],
        ["추천설정", "gender", "성별조건", "Gender", "추천매칭", "내부용", "필수", "male/female", "05_추천설정_팁주의", ""],
        ["추천설정", "experienceLevel", "경험레벨조건", "Experience", "추천매칭", "내부용", "필수", "beginner/intermediate/advanced", "05_추천설정_팁주의", ""],
        ["추천설정", "heightMinCm", "키최소", "Height Min", "추천매칭", "내부용", "필수", "숫자(cm)", "05_추천설정_팁주의", ""],
        ["추천설정", "heightMaxCm", "키최대", "Height Max", "추천매칭", "내부용", "필수", "숫자(cm)", "05_추천설정_팁주의", ""],
        ["추천팁", "tips.ko", "운동팁(한글)", "Workout Tips KO", "추천결과/휴식음성", "표시중", "권장", "줄바꿈 문장목록", "05_추천설정_팁주의", "TTS 핵심"],
        ["추천팁", "tips.en", "운동팁(영문)", "Workout Tips EN", "추천결과", "표시중", "권장", "줄바꿈 문장목록", "05_추천설정_팁주의", ""],
        ["추천팁", "warnings.ko", "주의(한글)", "Warnings KO", "추천결과/휴식음성", "표시중", "권장", "줄바꿈 문장목록", "05_추천설정_팁주의", "TTS 핵심"],
        ["추천팁", "warnings.en", "주의(영문)", "Warnings EN", "추천결과", "표시중", "권장", "줄바꿈 문장목록", "05_추천설정_팁주의", ""],
        # Video
        ["영상", "youtubeId", "유튜브ID", "YouTube ID", "추천결과(예정)", "데이터만", "선택", "문자열", "-", "현재 UI 없음"],
        ["영상", "videoUrl", "머신영상URL", "Video URL", "상세(예정)", "데이터만", "선택", "URL", "-", "현재 UI 없음"],
        # Gym machines
        ["헬스장머신", "machineCode", "머신코드", "Machine Code", "헬스장상세", "표시중", "필수", "문자열", "06_헬스장보유머신", ""],
        ["헬스장머신", "machineName", "머신표시명", "Machine Name", "헬스장상세/재고", "표시중", "필수", "문자열", "06_헬스장보유머신", ""],
        ["헬스장머신", "quantity", "대수", "Quantity", "헬스장상세", "표시중", "선택", "정수", "06_헬스장보유머신", "대"],
        ["헬스장머신", "floorZone", "위치구역", "Floor Zone", "헬스장상세", "표시중", "선택", "문자열", "06_헬스장보유머신", "예: 1층 하체존"],
        ["헬스장머신", "isAvailable", "사용가능", "Available", "헬스장상세", "표시중", "필수", "TRUE/FALSE", "06_헬스장보유머신", ""],
        ["헬스장머신", "notes", "메모", "Notes", "헬스장상세(비가용시)", "표시중", "선택", "문자열", "06_헬스장보유머신", ""],
        # User content / not catalog
        ["사용자", "personalTipMemo", "나만의 팁 메모장", "My tip notebook", "운동기록", "표시중", "선택", "사용자입력", "-", "카탈로그 작성대상 아님"],
        ["요청게시판", "brandName/machineName/description", "브랜드/기구명/설명", "Request fields", "머신요청게시판", "표시중", "부분필수", "자유텍스트", "-", "카탈로그 아님"],
    ]
    add_rows(ws, rows)
    for r in range(2, len(rows) + 2):
        status = ws.cell(r, 6).value
        if status == "표시중":
            ws.cell(r, 6).fill = PatternFill("solid", fgColor="DCFCE7")
        elif status == "데이터만":
            ws.cell(r, 6).fill = PatternFill("solid", fgColor="FEF9C3")
        elif status == "내부용":
            ws.cell(r, 6).fill = PatternFill("solid", fgColor="E2E8F0")
        if ws.cell(r, 7).value == "필수":
            ws.cell(r, 7).fill = REQ_FILL
    autosize(ws, {1: 12, 2: 34, 3: 22, 4: 22, 5: 34, 6: 12, 7: 10, 8: 24, 9: 20, 10: 28})


def sheet_brands(wb: Workbook):
    ws = wb.create_sheet("02_브랜드")
    headers = [
        "brandCode*",
        "slug*",
        "name_ko*",
        "name_en*",
        "name_ja",
        "name_zh",
        "description_ko",
        "description_en",
        "logoFile",
        "logoUrl",
        "websiteUrl",
        "isActive*",
        "비고",
    ]
    ws.append(headers)
    style_header(ws, len(headers))
    for c in range(1, 5):
        ws.cell(1, c).fill = PatternFill("solid", fgColor="B45309")
    sample = [
        "CYBEX",
        "cybex",
        "사이벡스",
        "Cybex",
        "サイベックス",
        "赛百斯",
        "사이벡스는 생체역학 기반의 셀렉터라이즈드 라인으로 잘 알려진 글로벌 피트니스 브랜드입니다. 안정적인 동작 궤적으로 초보자도 안전하게 힘을 쓰기 쉽습니다.",
        "Cybex is known for biomechanics-driven selectorized strength lines with stable movement paths.",
        "cybex.jpg",
        "",
        "https://www.cybexintl.com",
        "TRUE",
        "예시 행 — 삭제 후 작성 가능",
    ]
    add_rows(ws, [sample], fill=SAMPLE_FILL)
    # empty template rows
    add_rows(ws, [[""] * len(headers) for _ in range(10)], start_row=3)
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("L2:L200")
    autosize(ws, {1: 18, 2: 16, 3: 16, 4: 16, 5: 14, 6: 14, 7: 40, 8: 40, 9: 18, 10: 28, 11: 28, 12: 10, 13: 22})


def sheet_machines(wb: Workbook):
    ws = wb.create_sheet("03_머신")
    headers = [
        "machineCode*",
        "brandCode*",
        "name_ko*",
        "name_en*",
        "name_ja",
        "name_zh",
        "muscleGroup*",
        "machineType*",
        "description_ko",
        "description_en",
        "howTo_ko (줄바꿈)",
        "howTo_en (줄바꿈)",
        "warnings_ko (줄바꿈)",
        "warnings_en (줄바꿈)",
        "tips_ko (줄바꿈)",
        "tips_en (줄바꿈)",
        "beginnerTips_ko (줄바꿈)",
        "beginnerTips_en (줄바꿈)",
        "recommendedExperience",
        "hasSeat*",
        "hasBackPad*",
        "hasFootPlate*",
        "hasHandle*",
        "romType",
        "isActive*",
        "비고",
    ]
    ws.append(headers)
    style_header(ws, len(headers))
    sample = [
        "CY_LEG_CURL",
        "CYBEX",
        "레그 컬",
        "Leg Curl",
        "レッグカール",
        "腿部弯举",
        "legs",
        "selectorized",
        "레그 컬은 햄스트링을 중심으로 허벅지 뒤를 고립해서 단련하는 머신입니다.",
        "The leg curl isolates the hamstrings for controlled posterior-chain work.",
        "시트 높이를 무릎 피벗에 맞춥니다.\n발목 패드를 아킬레스 위쪽에 둡니다.\n허리를 패드에 고정한 뒤 시작합니다.\n숨을 내쉬며 발뒤꿈치를 엉덩이 쪽으로 당깁니다.\n수축 지점에서 잠깐 멈춥니다.\n천천히 이완하며 시작 자세로 돌아옵니다.\n무릎이 피벗에서 벗어나지 않게 유지합니다.\n반동 없이 일정한 템포로 반복합니다.\n통증이 있으면 즉시 멈춥니다.\n세트 후 가볍게 스트레칭합니다.",
        "Align the knee with the machine pivot.\nPlace the ankle pad above the Achilles.\nKeep hips pressed into the pad.\nExhale as you curl the heels toward the glutes.\nPause briefly at peak contraction.\nLower under control.\nKeep knees tracking with the pivot.\nAvoid bouncing.\nStop if pain appears.\nStretch lightly after the set.",
        "관절을 완전히 잠그지 않습니다.\n반동을 사용하지 않습니다.\n허리를 과도하게 꺾지 않습니다.\n엉덩이가 패드에서 뜨지 않게 합니다.\n무게를 급하게 올리지 않습니다.\n무릎이 안밖으로 흔들리지 않게 합니다.\n발목을 과도하게 꺾지 않습니다.\n가동범위를 억지로 늘리지 않습니다.\n워밍업 없이 고중량으로 시작하지 않습니다.\n통증이 있으면 즉시 중단합니다.",
        "Do not lock the joints hard.\nDo not use momentum.\nDo not over-arch the lower back.\nKeep hips on the pad.\nDo not jump the weight.\nKeep knees stable.\nAvoid extreme ankle bend.\nDo not force range of motion.\nWarm up before heavy sets.\nStop if pain occurs.",
        "수축 지점에서 1초 정지합니다.\n천천히 이완합니다.\n목표 근육의 긴장을 유지합니다.\n발끝을 살짝 모아 햄스트링 자극을 느낍니다.\n호흡을 끊지 않습니다.\n가동범위는 통증이 없는 범위로 잡습니다.\n세트 사이 호흡을 정리합니다.\n마지막 2회는 더 천천히 합니다.\n좌우 밸런스를 확인합니다.\n가벼운 중량으로 감각을 먼저 찾습니다.",
        "Pause one second at peak contraction.\nLower slowly.\nKeep tension on the target muscle.\nSlightly point toes to feel hamstrings.\nKeep breathing.\nUse a pain-free range.\nReset breathing between sets.\nSlow the last two reps.\nCheck left-right balance.\nStart light to find the feel.",
        "처음에는 가벼운 중량으로 10~12회부터 시작합니다.\n시트와 패드 위치를 먼저 맞춥니다.\n거울보다 근육 감각에 집중합니다.\n무리한 가동범위는 피합니다.\n세트는 2~3세트면 충분합니다.\n통증이 있으면 트레이너에게 알립니다.\n운동 전 가벼운 워밍업을 합니다.\n호흡을 참고 하지 않습니다.\n혼자 고중량을 시도하지 않습니다.\n끝나면 허벅지 뒤를 가볍게 풀어줍니다.",
        "Start light for 10–12 reps.\nSet seat and pads first.\nFocus on muscle feel, not mirrors.\nAvoid forced range.\nTwo to three sets is enough.\nTell a trainer if pain appears.\nWarm up first.\nDo not hold your breath.\nDo not jump to heavy loads alone.\nStretch hamstrings after.",
        "beginner",
        "TRUE",
        "TRUE",
        "TRUE",
        "FALSE",
        "fixed",
        "TRUE",
        "예시 — 레그 컬",
    ]
    add_rows(ws, [sample], fill=SAMPLE_FILL)
    add_rows(ws, [[""] * len(headers) for _ in range(15)], start_row=3)

    for col, formula in {
        "G": '"chest,back,legs,shoulders,biceps,triceps,core,full_body"',
        "H": '"plate_loaded,selectorized,cable,smith,free_weight,bodyweight"',
        "S": '"beginner,intermediate,advanced"',
        "T": '"TRUE,FALSE"',
        "U": '"TRUE,FALSE"',
        "V": '"TRUE,FALSE"',
        "W": '"TRUE,FALSE"',
        "Y": '"TRUE,FALSE"',
    }.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}500")

    autosize(
        ws,
        {
            1: 18,
            2: 16,
            3: 16,
            4: 18,
            5: 14,
            6: 14,
            7: 14,
            8: 14,
            9: 28,
            10: 28,
            11: 36,
            12: 36,
            13: 36,
            14: 36,
            15: 36,
            16: 36,
            17: 36,
            18: 36,
            19: 16,
            20: 10,
            21: 12,
            22: 12,
            23: 10,
            24: 10,
            25: 10,
            26: 16,
        },
    )


def sheet_images(wb: Workbook):
    ws = wb.create_sheet("04_머신이미지")
    headers = [
        "machineCode*",
        "imageFile",
        "imageUrl",
        "thumbnailFile",
        "altText_ko",
        "altText_en",
        "sortOrder*",
        "isPrimary*",
        "비고",
    ]
    ws.append(headers)
    style_header(ws, len(headers))
    samples = [
        [
            "CY_LEG_CURL",
            "cy_leg_curl.jpg",
            "",
            "cy_leg_curl_thumb.jpg",
            "사이벡스 레그 컬 머신",
            "Cybex leg curl machine",
            0,
            "TRUE",
            "대표 이미지",
        ],
        [
            "CY_LEG_CURL",
            "cy_leg_curl_side.jpg",
            "",
            "",
            "사이드 뷰",
            "Side view",
            1,
            "FALSE",
            "추가 갤러리(예정)",
        ],
    ]
    add_rows(ws, samples, fill=SAMPLE_FILL)
    add_rows(ws, [[""] * len(headers) for _ in range(20)], start_row=4)
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("H2:H500")
    autosize(ws, {1: 18, 2: 24, 3: 36, 4: 24, 5: 24, 6: 24, 7: 12, 8: 12, 9: 18})


def sheet_settings(wb: Workbook):
    ws = wb.create_sheet("05_추천설정_팁주의")
    headers = [
        "machineCode*",
        "gender*",
        "experienceLevel*",
        "heightMinCm*",
        "heightMaxCm*",
        "seatPosition",
        "backPadPosition",
        "footPosition",
        "handlePosition",
        "romSetting",
        "recommendedWeightKg",
        "recommendedRepsMin",
        "recommendedRepsMax",
        "tips_ko (줄바꿈)*",
        "tips_en (줄바꿈)",
        "warnings_ko (줄바꿈)*",
        "warnings_en (줄바꿈)",
        "비고",
    ]
    ws.append(headers)
    style_header(ws, len(headers))
    sample = [
        "CY_LEG_CURL",
        "male",
        "intermediate",
        165,
        185,
        5,
        3,
        4,
        "",
        "최대",
        40,
        10,
        15,
        "엉덩이를 패드에 고정합니다.\n전 가동범위를 사용합니다.\n수축 지점에서 짧게 멈춥니다.\n천천히 이완합니다.\n호흡을 유지합니다.\n무릎을 머신 피벗에 맞춥니다.\n반동 없이 당깁니다.\n마지막 반복은 더 천천히 합니다.\n세트 사이 호흡을 정리합니다.\n가벼운 세트로 워밍업합니다.",
        "Keep hips pressed into the pad.\nUse a full pain-free range.\nPause briefly at peak contraction.\nLower slowly.\nKeep breathing.\nAlign knees with the pivot.\nCurl without momentum.\nSlow the last reps.\nReset breathing between sets.\nWarm up light.",
        "엉덩이를 들지 마세요.\n허리를 과도하게 꺾지 마세요.\n관절을 완전히 잠그지 마세요.\n무게를 급하게 당기지 마세요.\n무릎이 흔들리지 않게 하세요.\n통증이 있으면 즉시 멈추세요.\n워밍업 없이 고중량을 쓰지 마세요.\n발목을 과도하게 꺾지 마세요.\n가동범위를 억지로 늘리지 마세요.\n세트 중 숨을 참고 있지 마세요.",
        "Do not lift the hips.\nDo not over-arch the back.\nDo not lock joints hard.\nDo not yank the weight.\nKeep knees stable.\nStop if pain occurs.\nDo not skip warm-up.\nAvoid extreme ankle bend.\nDo not force ROM.\nDo not hold your breath.",
        "추천결과·휴식 음성에 실제로 사용",
    ]
    add_rows(ws, [sample], fill=SAMPLE_FILL)
    add_rows(ws, [[""] * len(headers) for _ in range(15)], start_row=3)
    for col, formula in {
        "B": '"male,female"',
        "C": '"beginner,intermediate,advanced"',
    }.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}500")
    autosize(
        ws,
        {
            1: 16,
            2: 10,
            3: 16,
            4: 12,
            5: 12,
            6: 12,
            7: 14,
            8: 12,
            9: 12,
            10: 12,
            11: 14,
            12: 12,
            13: 12,
            14: 36,
            15: 36,
            16: 36,
            17: 36,
            18: 24,
        },
    )


def sheet_gym(wb: Workbook):
    ws = wb.create_sheet("06_헬스장보유머신")
    headers = [
        "gymSlug_or_Id",
        "machineCode*",
        "machineName_ko",
        "quantity",
        "floorZone",
        "isAvailable*",
        "notes",
        "비고",
    ]
    ws.append(headers)
    style_header(ws, len(headers))
    sample = [
        "gangnam-strength",
        "CY_LEG_CURL",
        "레그 컬",
        2,
        "1층 하체존",
        "TRUE",
        "",
        "헬스장 상세 인벤토리용(선택 시트)",
    ]
    add_rows(ws, [sample], fill=SAMPLE_FILL)
    add_rows(ws, [[""] * len(headers) for _ in range(10)], start_row=3)
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("F2:F200")
    autosize(ws, {1: 18, 2: 16, 3: 16, 4: 10, 5: 16, 6: 12, 7: 28, 8: 28})


def sheet_example(wb: Workbook):
    ws = wb.create_sheet("07_작성예시")
    ws["A1"] = "한 기구를 완전히 채울 때 작성 순서"
    ws["A1"].font = Font(bold=True, size=14)
    steps = [
        "1) 02_브랜드에 브랜드 1행 작성",
        "2) 03_머신에 기구 1행 작성 (사용방법/주의/팁/초보팁 각 10줄 권장)",
        "3) 04_머신이미지에 대표이미지 1행(isPrimary=TRUE) + 필요시 추가컷",
        "4) 05_추천설정_팁주의에 성별/경험/키구간별 추천값 + 운동팁/주의 작성",
        "5) (선택) 06_헬스장보유머신에 실제 보유 헬스장 정보",
        "",
        "이미지 파일은 별도 폴더로 함께 전달해 주세요.",
        "권장 경로 예: assets/brands/{slug}.jpg , assets/machines/{slug}/{machine_code_lower}.jpg",
        "예: assets/machines/cybex/cy_leg_curl.jpg",
        "",
        "현재 서비스에서 바로 보이는 핵심 항목",
        "- 브랜드: 이름, 소개, 로고, 홈페이지",
        "- 머신 목록/상세: 이름, 브랜드명, 운동부위, 머신종류, 대표이미지",
        "- 추천결과: 시트/백패드/발/핸들/가동범위/추천중량/추천횟수, 운동팁, 주의",
        "",
        "지금은 화면에 안 보이지만 미리 채워 두면 좋은 항목",
        "- 사용방법(howTo), 초보자팁, 머신소개, 추천경험레벨, 추가 이미지/유튜브",
    ]
    for i, line in enumerate(steps, start=3):
        ws.cell(i, 1, line).alignment = WRAP
    ws.column_dimensions["A"].width = 100


def main():
    wb = Workbook()
    sheet_guide(wb)
    sheet_dictionary(wb)
    sheet_brands(wb)
    sheet_machines(wb)
    sheet_images(wb)
    sheet_settings(wb)
    sheet_gym(wb)
    sheet_example(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT_COPY.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    wb.save(OUT_COPY)
    print(f"Wrote {OUT}")
    print(f"Wrote {OUT_COPY}")


if __name__ == "__main__":
    main()
